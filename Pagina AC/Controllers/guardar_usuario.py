import openpyxl
from datetime import datetime

def guardar_en_excel(nombre_usuario):
    # Crear un nuevo libro de trabajo o cargar uno existente
    try:
        # Intenta cargar un archivo existente
        workbook = openpyxl.load_workbook('usuarios.xlsx')
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()

    # Selecciona la hoja activa
    sheet = workbook.active

    # Agregar encabezados si la hoja está vacía
    if sheet.max_row == 1:
        sheet['A1'] = 'Usuario'
        sheet['B1'] = 'Fecha y Hora'

    # Obtener la fecha y hora actual
    fecha_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Agregar la información del usuario a la hoja
    sheet.append([nombre_usuario, fecha_hora])

    # Guardar el libro de trabajo
    workbook.save('usuarios.xlsx')

if __name__ == '__main__':
    # Aquí puedes ingresar el nombre de usuario que quieres guardar
    nombre_usuario = input("Ingresa el nombre de usuario: ")
    guardar_en_excel(nombre_usuario)
    print("Información guardada en usuarios.xlsx")
